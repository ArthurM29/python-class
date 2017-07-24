import openpyxl

file = '/Users/amanasyan/Documents/Python projects/python-class/Advanced level/Excel/example.xlsx'
# open 'Excel' object and return file object
wb = openpyxl.load_workbook(file)

# return a list of sheet names
sheet_names = wb.get_sheet_names()

# get specific sheet name - returns 'Worksheet' object
sheet = wb.get_sheet_by_name('Sheet1')

# get title of the sheet
sheet_title = sheet.title

# get the active sheet
active_sheet = wb.active

# get one cell from sheet
cell = active_sheet['A1']

# and its row, column and value
cell_row = cell.row
cell_column = cell.column
cell_value = cell.value
coordinate = cell.coordinate

# access columns by numbers instead of letters
# print(active_sheet.max_row)
# print(active_sheet.max_column)

# get cells of a rectangular slice
# for row in sheet['A1':'C3']:
#     for cellObj in row:
#         print(cellObj.coordinate, cellObj.value)
#     print('--- END OF ROW ---')

columns = list(sheet.columns)
for cell in columns[1]:
	print(cell.value)




















